"""
إنشاء جدول Excel من البيانات المستخرجة
"""

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def create_excel(data_list, output_file="report.xlsx"):
    """
    إنشاء ملف Excel من قائمة البيانات
    
    Args:
        data_list: قائمة من القواميس (كل قاموس = صف واحد)
        output_file: اسم ملف الإخراج
    
    Returns:
        str: مسار الملف المُنشأ
    """
    
    # تحويل القائمة إلى DataFrame
    df = pd.DataFrame(data_list)
    
    # ترتيب الأعمدة
    columns_order = [
        'رقم_الأمر',
        'التاريخ',
        'اسم_الشركة',
        'البيان',
        'الجهة_الطالبة',
        'المبلغ_الإجمالي'
    ]
    
    # التأكد من وجود كل الأعمدة
    for col in columns_order:
        if col not in df.columns:
            df[col] = ''
    
    df = df[columns_order]
    
    # تسمية الأعمدة بالعربي
    df.columns = [
        'رقم الأمر',
        'التاريخ',
        'اسم الشركة',
        'البيان',
        'الجهة الطالبة',
        'المبلغ الإجمالي'
    ]
    
    # حفظ في Excel
    df.to_excel(output_file, index=False, engine='openpyxl')
    
    # تنسيق الملف
    format_excel(output_file)
    
    print(f"✅ تم إنشاء الملف: {output_file}")
    return output_file


def format_excel(file_path):
    """
    تنسيق ملف Excel ليبدو احترافياً
    """
    # فتح الملف
    wb = load_workbook(file_path)
    ws = wb.active
    
    # ألوان
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # حدود
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # تنسيق الهيدر (الصف الأول)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # تنسيق باقي الصفوف
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            cell.border = thin_border
    
    # ضبط عرض الأعمدة
    column_widths = {
        'A': 10,  # رقم الأمر
        'B': 12,  # التاريخ
        'C': 30,  # اسم الشركة
        'D': 50,  # البيان
        'E': 25,  # الجهة الطالبة
        'F': 15   # المبلغ الإجمالي
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # ضبط ارتفاع الصفوف
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 40
    
    # حفظ التغييرات
    wb.save(file_path)


# في excel_generator.py

def add_to_existing_excel(data_list, excel_file="monthly_report.xlsx"):
    """
    إضافة بيانات جديدة إلى ملف Excel موجود
    أو إنشاء ملف جديد إذا لم يكن موجوداً
    """
    try:
        # محاولة قراءة الملف الموجود
        df_existing = pd.read_excel(excel_file, engine='openpyxl') # إضافة engine لتحسين التوافق
        
        # إضافة البيانات الجديدة (باستخدام data_list)
        df_new = pd.DataFrame(data_list)
        df_combined = pd.concat([df_existing, df_new], ignore_index=True)
        
        # حفظ الملف
        df_combined.to_excel(excel_file, index=False, engine='openpyxl')
        format_excel(excel_file)
        
        print(f"✅ تمت إضافة البيانات إلى: {excel_file}")
        
    except FileNotFoundError:
        # إنشاء ملف جديد
        create_excel(data_list, excel_file)
    except Exception as e:
        print(f"❌ خطأ أثناء إضافة البيانات لملف Excel موجود: {e}")


# للاختبار
if __name__ == "__main__":    # بيانات تجريبية
    test_data = [
        {
            'رقم_الأمر': '123',
            'التاريخ': '1/10/2025',
            'اسم_الشركة': 'شيخ لبيع إطارات وبطاريات السيارات',
            'البيان': 'توريد عدد (8) إطار كاوتش...',
            'الجهة_الطالبة': 'قطاع الوقاية جنوب',
            'المبلغ_الإجمالي': '20800.00'
        },
        {
            'رقم_الأمر': '124',
            'التاريخ': '1/10/2025',
            'اسم_الشركة': 'عالم البطاريات والاطارات',
            'البيان': 'توريد عدد (2) بطارية جافة...',
            'الجهة_الطالبة': 'قطاع الوقاية شمال وشرق',
            'المبلغ_الإجمالي': '5630.00'
        }
    ]
    
    print("🔨 جاري إنشاء ملف Excel...")
    create_excel(test_data, "test_report.xlsx")